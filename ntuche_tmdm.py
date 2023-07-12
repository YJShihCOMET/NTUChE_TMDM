import numpy as np
import pandas as pd
import os
import openpyxl
import decimal
from openpyxl.utils.dataframe import dataframe_to_rows

class arrangement:
    """
    Parameters
    ----------
    grade_dict: dictionary, str:float
        將等第成績轉換為等第積分的字典
    """
    
    grade_dict = {
        'A+':4.3, 
        'A':4.0, 
        'A-':3.7, 
        'B+':3.3, 
        'B':3.0, 
        'B-':2.7, 
        'C+':2.3, 
        'C':2.0,
        'C-':1.7, 
        'F':0.0
                 }
    
    def __init__(self, grade_path, core_course1):
        """
        
        初始化
        
        ----------
        Parameters
        ----------
        grade_path : str
            學生成績的檔案路徑
        __df_alldata : pd.DataFrame
            所有學生的所有平均分數資料
        """
        self.grade_path = grade_path
        self.core_course1 = core_course1
        self.__df_alldata = None
    
    @staticmethod
    def modify_round(x, dec=2):
        """
        
        四捨五入的函數
        (python的round、numpy的round和around都常常會回傳錯誤的結果, eg. 4.165回傳4.16給我)
        
        ----------
        Parameters
        ----------
        x: float
            想要取四捨五入的數值
        dec: float
            四捨五入的精確度(eg.想要取至小數點下第2位就設為2)
        rounded_x: str/float
            取完四捨五入以後的數值
        """
        x_str = str(x)
        x = decimal.Decimal(x_str)
        rounded_x = x.quantize(decimal.Decimal(str(10**(-dec))), rounding=decimal.ROUND_HALF_UP)
        return float(rounded_x)
    
    @staticmethod
    def dedupe(items):
        """
        
        在不影響順序的情況下刪除重複元素的函數
        
        ----------
        Parameters
        ----------
        items: iterable list, int
            有許多重複元素的list或array
        """
        seen = set()
        for item in items:
            if item not in seen:
                yield item
                seen.add(item)
    
    @property
    def df_gradedata(self): #學生成績
        """
        
        將grade_path路徑中的學生成績轉為DataFrame並刪除重複的成績，呼叫此函數即可獲得所有學生所有成績的總表
        
        ----------
        Parameters
        ----------
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        col: list, str
            成績總表的欄位名稱
        drop_index_list: list, int
            含有所有重複成績的索引列表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        """
        df_gradedata = pd.read_excel(self.grade_path).replace('\xa0\xa0', np.nan)
        col = df_gradedata.iloc[1].to_list() # 取得欄位名
        for i, coli in enumerate(col): # 若欄位名中有名為"課號"的欄，將其改為課程識別碼
            if coli == '課號':
                col[i] = '課程識別碼'
        df_gradedata.columns = col
        df_gradedata = df_gradedata.iloc[2:].reset_index(drop=True) # 刪除前兩列並重設index
        drop_index_list = []
        sheetname = []
        for i in range(1, len(df_gradedata)): # 刪除重複成績並獲得每學期的名稱
            if ( df_gradedata['學年'].iloc[i] == df_gradedata['學年'].iloc[i-1] ) and ( df_gradedata['學期'].iloc[i] == df_gradedata['學期'].iloc[i-1] )\
             and ( df_gradedata['學號'].iloc[i] == df_gradedata['學號'].iloc[i-1] ) and ( df_gradedata['課程識別碼'].iloc[i] == df_gradedata['課程識別碼'].iloc[i-1] )\
             and ( df_gradedata['學分'].iloc[i] == df_gradedata['學分'].iloc[i-1] ):
                drop_index_list.append(i)
            sheetnamei = str(df_gradedata['學年'].iloc[i]) + '_' + str(df_gradedata['學期'].iloc[i])
            if sheetnamei not in sheetname:
                sheetname.append(sheetnamei)
        df_gradedata = df_gradedata.drop(index=drop_index_list).reset_index(drop=True)
        return df_gradedata, sheetname
    
    @property
    def df_gradedata_split(self): #學生成績
        """
        
        將df_gradedata中的成績總表分割為不同學期的成績表
        
        ----------
        Parameters
        ----------
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        """
        df_gradedata, sheetname = self.df_gradedata
        df_gradedata_split = []
        for sheetnamei in sheetname:
            year, semester = sheetnamei.split('_')
            df_gradedata_spliti = df_gradedata.loc[(df_gradedata['學年'] == int(year)) & (df_gradedata['學期'] == int(semester))]
            df_gradedata_split.append(df_gradedata_spliti.reset_index(drop=True))
        return df_gradedata_split, sheetname
    
    @property
    def all_students_id_split(self): 
        """
        Parameters
        ----------
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_students_id_split: 2d list, str
            不同學期所有學生的學號
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_id_split = []
        for df_gradedata_spliti in df_gradedata_split:
            all_students_id_split.append([i.strip() for i in self.dedupe((df_gradedata_spliti['學號']))])
        return all_students_id_split
    
    @property
    def all_students_id(self): 
        """
        Parameters
        ----------
        all_students_id_split: 2d list, str
            不同學期所有學生的學號
        all_students_id: list, str
            所有學生的學號
        """
        all_students_id_split = self.all_students_id_split
        all_students_id = list(self.dedupe(sum(all_students_id_split, [])))
        return all_students_id
    
    @property
    def all_students_name_split(self):
        """
        Parameters
        ----------
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_students_name_split: 2d list, str
            不同學期所有學生的姓名
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_name_split = []
        for df_gradedata_spliti in df_gradedata_split:
            all_students_name_split.append([i.strip() for i in self.dedupe((df_gradedata_spliti['學生姓名']))])
        return all_students_name_split
    
    @property
    def all_students_name(self):
        """
        Parameters
        ----------
        all_students_name_split: 2d list, str
            不同學期所有學生的姓名
        all_students_name: list, str
            所有學生的姓名
        """
        all_students_name_split = self.all_students_name_split
        all_students_name = list(self.dedupe(sum(all_students_name_split, [])))
        return all_students_name
    
    @property
    def all_students_year_split(self):
        """
        Parameters
        ----------
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_students_year_split: 2d list, str
            不同學期所有學生的年級
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_id_split = self.all_students_id_split
        all_students_year_split = []
        for df_gradedata_spliti, all_students_idi in zip(df_gradedata_split, all_students_id_split):
            all_students_year = []
            for students_id in all_students_idi:
                try:
                    students_year = df_gradedata_spliti[df_gradedata_spliti['學號']==students_id]['年級'].iloc[0]
                except:
                    students_year = np.nan
                all_students_year.append(students_year)
            all_students_year_split.append(all_students_year)
        return all_students_year_split
    
    @property
    def all_students_year(self):
        """
        Parameters
        ----------
        all_students_name_split: 2d list, str
            不同學期所有學生的姓名
        all_students_year: list, str
            所有學生的年級
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_id = self.all_students_id
        all_students_year = []
        for students_id in all_students_id:
            students_year = None
            for df_gradedata_spliti in df_gradedata_split:
                try:
                    students_year = df_gradedata_spliti[df_gradedata_spliti['學號']==students_id]['年級'].iloc[0]
                except:
                    students_year = None
                if students_year is not None:
                    all_students_year.append(students_year)
                    break
        return all_students_year
    
    @property
    def all_students_department_split(self):
        """
        Parameters
        ----------
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_students_id_split: 2d list, str
            不同學期所有學生的學號
        all_students_department_split: 2d list, str
            不同學期所有學生的系所名稱
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_id_split = self.all_students_id_split
        all_students_department_split = []
        for df_gradedata_spliti, all_students_idi in zip(df_gradedata_split, all_students_id_split):
            all_students_department = []
            for students_id in all_students_idi:
                try:
                    students_department = df_gradedata_spliti[df_gradedata_spliti['學號']==students_id]['學生本學系'].iloc[0].strip()
                except:
                    students_department = np.nan
                all_students_department.append(students_department)
            all_students_department_split.append(all_students_department)
        return all_students_department_split
    
    @property
    def all_students_department(self):
        """
        Parameters
        ----------
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_students_id: list, str
            所有學生的學號
        all_students_department: list, str
            所有學生的系所名稱
        """
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_students_id = self.all_students_id
        all_students_department = []
        for students_id in all_students_id:
            students_department = None
            for df_gradedata_spliti in df_gradedata_split:
                try:
                    students_department = df_gradedata_spliti[df_gradedata_spliti['學號']==students_id]['學生本學系'].iloc[0].strip()
                except:
                    students_department = None
                if students_department is not None:
                    all_students_department.append(students_department)
                    break
        return all_students_department
    
    def calc_allavg(self, student_id, full_output=False):
        """
        
        計算一個學生每個學期所有科目的總平均
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        full_output: boolean
            是否需要輸出學生的總學分數
        grade_dict: dictionary, str:float
            將等第成績轉換為等第積分的字典
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        allcredits: np.array, int
            不同學期學生的總學分數
        allavgs: np.array, float
            不同學期學生所有科目的平均分數
        allcredit: np.array, int
            學生的總學分數
        allavg: np.array, float
            學生所有科目的平均分數
        grade: np.array, float
            學生各科的成績(等第積分)
        credit: np.array, int
            學生各科的學分數
        flag: boolean
            是否已搜尋到此學生的資料
        """
        grade_dict = self.grade_dict
        df_gradedata_split, _ = self.df_gradedata_split
        allcredits = np.array([])
        allavgs = np.array([])
        for df_gradedata_spliti in df_gradedata_split:
            grade = np.array([])
            credit = np.array([])
            flag = False
            for (ide, cde, gne) in zip(df_gradedata_spliti['學號'], df_gradedata_spliti['學分'], df_gradedata_spliti['成績']):
                if ( ide.strip() == student_id ) and ( type(gne) is str ):
                    grade = np.append(grade, grade_dict[gne.strip()])
                    credit = np.append(credit, cde)
                    flag = True
                if ide.strip() != student_id and flag:
                    break
            allcredit = sum(credit)
            allavg = np.sum( grade * credit ) / allcredit if ( grade.size != 0 ) and ( allcredit != 0 ) else 0
            allavg = self.modify_round(allavg)
            allcredits = np.append(allcredits, allcredit)
            allavgs = np.append(allavgs, allavg)
        if full_output:
            return allavgs, allcredits
        else:
            return allavgs
    
    def calc_core1avg(self, student_id, full_output=False):
        """
        
        計算一個學生的微積分、普通物理學與普通化學的三科平均
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        full_output: boolean
            是否需要輸出學生的各個必修課目的課程名稱、等第成績、等第積分與學分數資料
        grade_dict: dictionary, str:float
            將等第成績轉換為等第積分的字典
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        core_course1 : list, str
            本校所有微積分、普通化學與普通物理學課名的共通字串
        core_course1_name: list, str
            學生修習的微積分、普通化學或普通物理學課名
        grade: np.array, float
            學生修習的微積分、普通化學或普通物理學成績(等第積分)
        credit: np.array, int
            學生修習的微積分、普通化學或普通物理學的學分數
        gdcddata: list, str
            學生修習的微積分、普通化學或普通物理學的"等第成績 等第積分 學分數"
        flag: boolean
            是否已搜尋到此學生的資料
        core1credit: int
            學生修習微積分、普通化學或普通物理學的總學分數
        core1avg: float
            學生修習的微積分、普通化學或普通物理學的平均分數
        fulldata: dict, tuple, str
            包含學生修習的微積分、普通化學或普通物理學的課程名稱、等第成績、等第積分與學分數, 等第成績代表A+, A, A-, ...等
        """
        grade_dict = self.grade_dict
        df_gradedata_split, sheetname = self.df_gradedata_split
        core_course1 = self.core_course1
        core_course1_name = []
        grade = np.array([])
        credit = np.array([])
        gdcddata = []
        for df_gradedata_spliti, sheetnamei in zip(df_gradedata_split, sheetname):
            flag = False
            for (ide, cne, cde, gne) in zip(df_gradedata_spliti['學號'], df_gradedata_spliti['課名'], df_gradedata_spliti['學分'], df_gradedata_spliti['成績']):
                # print(cne, type(cne))
                if ( ide.strip() == student_id ) and ( type(gne) is str ) and ( type(cne) is str ):
                    if ( ( core_course1[0] in cne ) or ( core_course1[1] in cne ) or ( core_course1[2] in cne ) ) and ( '實驗' not in cne ) and ( cne not in core_course1_name ): # 若重複修習相同課名, 取最新的資料
                        core_course1_name.append(cne.strip())
                        gde = grade_dict[gne.strip()]
                        grade = np.append(grade, gde)
                        credit = np.append(credit, cde)
                        gdcddata.append(gne.strip() + ' ' + str(gde).strip() + ' ' + str(cde).strip())
                        flag = True
                if ide.strip() != student_id and flag:
                    break
        core1credit = sum(credit)
        core1avg = np.sum( grade * credit ) / core1credit if ( grade.size != 0 ) and ( core1credit != 0 ) else 0
        core1avg = self.modify_round(core1avg)
        if full_output:
            fulldata = dict(zip(core_course1_name, gdcddata))
            return core1avg, fulldata
        else:
            return core1avg
    
    
    def get_df_alldata(self):
        """
        
        獲得所有學生的所有平均分數以及修習的三科資料總表
        
        ----------
        Parameters
        ----------
        all_students_id: list, str
            所有學生的學號
        all_students_name: list, str
            所有學生的姓名
        all_students_department: list, str
            所有學生的系所名稱
        all_students_year: list, str
            所有學生的年級
        df_gradedata_split: list, pd.DataFrame
            不同學期所有學生的成績表
        sheetname: list, str
            含有所有學期名稱的列表 eg. 110_1(代表110學年度第一學期)
        all_allavg: list, np.array, float or 2d np.array, float
            所有學生每學期的的全科目平均分數
        all_allcredit: list, np.array, int or 2d np.array, int
            所有學生每學期的的總學分數
        all_core1avg: list, float
            所有學生的微積分、普通化學與普通物理學平均分數
        df_corse1data: pd.DataFrame
            所有學生修習的微積分、普通化學與普通物理學課程的等第成績、等第積分與學分數總表
        sheetname_new: list, str
            含有所有學期名稱平均和總學分數名稱的列表，為df_alldata中一部分的欄位名稱
        column: list, str
            df_avgdata的欄位名稱
        data_allavg_allcredit: list, np.array, (float or int)
            df_alldata中的所有學生的每學期所有科目平均與總學分數資料
        data: list, np.array, (float or int)
            df_alldata中的所有學生的學號、名字、系所名稱、每學期所有科目平均與總學分數以及三科平均的資料
        df_avgdata: pd.DataFrame
            所有學生的學號、名字、系所名稱、每學期所有科目平均與總學分數以及三科平均的總表
        df_alldata: pd.DataFrame
            df_avgdata和df_corse1data合併後的總表，其為所有學生的所有平均分數以及修習的三科資料總表
        """
        all_students_id = self.all_students_id
        all_students_name = self.all_students_name
        all_students_department = self.all_students_department
        all_students_year = self.all_students_year
        df_gradedata_split, sheetname = self.df_gradedata_split
        all_allavg = []
        all_allcredit = []
        all_core1avg = []
        df_corse1data = pd.DataFrame()
        sheetname_new = []
        for sheetnamei in sheetname:
            sheetname_new.append(sheetnamei+' 所有科目平均')
            sheetname_new.append(sheetnamei+' 總學分數')
        column = ['學號','學生姓名','學生本學系', '年級'] + sheetname_new + ['三科平均']
        for student_id in all_students_id: # 獲得所有學生成績資料的list
            allavg, allcredit = self.calc_allavg(student_id, True)
            core1avg, fulldata = self.calc_core1avg(student_id, True)
            all_allavg.append(allavg)
            all_allcredit.append(allcredit)
            df_corse1data = pd.concat([df_corse1data, pd.DataFrame([fulldata])], ignore_index=True) # 合併所有學生修習三科的成績資料
            all_core1avg.append(core1avg)
        
        # 獲得df_alldata表的資料
        data_allavg_allcredit = []
        all_allavg = np.array(all_allavg)
        all_allcredit = np.array(all_allcredit)
        for all_allavgi, all_allcrediti in zip(all_allavg.T, all_allcredit.T):
            data_allavg_allcredit.append(all_allavgi.tolist())
            data_allavg_allcredit.append(all_allcrediti.tolist())
        data = [all_students_id, all_students_name, all_students_department, all_students_year] +\
        data_allavg_allcredit + [all_core1avg]
        df_avgdata = pd.DataFrame(zip(*data), columns=column)
        df_alldata = pd.concat([df_avgdata, df_corse1data], axis=1)
        self.__df_alldata = df_alldata
        return df_alldata
    
    @property
    def df_alldata(self):
        """
        Parameters
        ----------
        df_alldata: pd.DataFrame
            所有學生的所有平均分數資料總表
        """
        if self.__df_alldata is None:
            return self.get_df_alldata()
        else:
            return self.__df_alldata
    
    @property
    def df_rankdata(self):
        """
        
        進行排名的計算(可有可無)
        
        ----------
        Parameters
        ----------
        col_all: list, str
            排名所依照的各種先後順序，順序由左到右
        df_alldata: pd.DataFrame
            所有學生的所有平均分數資料總表
        df_rankdata: pd.DataFrame
            包含所有學生所有平均分數資料的排名總表
        """
        col_all = ['三科平均']
        df_alldata = self.df_alldata
        df_rankdata = df_alldata.copy()
        ranklist = df_rankdata[col_all].apply(tuple, axis=1).rank(method='min', ascending=0)
        df_rankdata.insert(0, '排名', ranklist) #插入一欄紀錄每位學生的排名
        df_rankdata.sort_values('排名', inplace=True) #將此表格以排名來排序
        df_rankdata.index = df_alldata.index
        return df_rankdata
    
    def save_df_data(self, df_data, savepath, sheet_name, method='dataframe_to_rows'):
        """
        
        將排名後的資料儲存至指定路徑
        
        ----------
        Parameters
        ----------
        savepath: str
            排名結果的檔案儲存路徑
        sheet_name: str
            設定結果的excel檔中的工作表名稱
        method: str
            存檔的方法有兩種:
            1. ExcelWriter: 程式碼比較簡潔，但我當初在編寫時有時候會產生出損毀過的excel檔
            2. dataframe_to_rows: 程式碼看起來比較繁雜，但是可以產生出正常的excel檔
        df_rankdata: pd.DataFrame
            包含所有學生所有平均分數資料的排名總表
        """
        if method == 'ExcelWriter':
            if os.path.exists(savepath):
                writer = pd.ExcelWriter(savepath, engine='openpyxl', mode='a')
                book = openpyxl.load_workbook(savepath)
                writer.book = book
            else:
                writer = pd.ExcelWriter(savepath, engine='openpyxl')
                book = openpyxl.Workbook()
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
            writer.close()
        elif method == 'dataframe_to_rows':
            rows = dataframe_to_rows(df_data, index=False)
            if os.path.exists(savepath):
                book = openpyxl.load_workbook(savepath)
                sheet = book.create_sheet(title=sheet_name)
            else:
                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = sheet_name
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                     sheet.cell(row=r_idx, column=c_idx, value=value)
            book.save(filename=savepath)
        else:
            print('Please input "dataframe_to_rows" or "ExcelWriter" to method variable.')